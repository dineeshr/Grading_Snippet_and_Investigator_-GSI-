from curses.ascii import isalpha, isdigit
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb=Workbook()

print("************************************")
print("* GRADING SNIPPET AND INVESTIGATOR *")
print("************************************")

def gsi(rn,db,sheetname,file_name):

        print_sheetname=sheetname
        fn=str(file_name)+".xlsx"
        ro=1
        count=0
        la11apassed=la11afailed=la11apresent=la11aabsent=0
        cle1epassed=cle1efailed=cle1epresent=cle1eabsent=0
        lk11apassed=lk11afailed=lk11apresent=lk11aabsent=0
        lz11apassed=lz11afailed=lz11apresent=lz11aabsent=0
        se21apassed=se21afailed=se21apresent=se21aabsent=0
        se211passed=se211failed=se211present=se211absent=0
        sm3aapassed=sm3aafailed=sm3aapresent=sm3aaabsent=0
        nlt1cpassed=nlt1cfailed=nlt1cpresent=nlt1cabsent=0
        tlt1cpassed=tlt1cfailed=tlt1cpresent=tlt1cabsent=0
        sm5aapassed=sm5aafailed=sm5aapresent=sm5aaabsent=0
        pz1sapassed=pz1safailed=pz1sapresent=pz1saabsent=0
        la12apassed=la12afailed=la12apresent=la12aabsent=0
        cle2gpassed=cle2gfailed=cle2gpresent=cle2gabsent=0
        clk2tpassed=clk2tfailed=clk2tpresent=clk2tabsent=0
        lz12apassed=lz12afailed=lz12apresent=lz12aabsent=0
        su22apassed=su22afailed=su22apresent=su22aabsent=0
        su221passed=su221failed=su221present=su221absent=0
        sm3aepassed=sm3aefailed=sm3aepresent=sm3aeabsent=0
        nlt2dpassed=nlt2dfailed=nlt2dpresent=nlt2dabsent=0
        tlt2dpassed=tlt2dfailed=tlt2dpresent=tlt2dabsent=0
        sm5abpassed=sm5abfailed=sm5abpresent=sm5ababsent=0
        pz1scpassed=pz1scfailed=pz1scpresent=pz1scabsent=0
        sz23apassed=sz23afailed=sz23apresent=sz23aabsent=0
        sz23bpassed=sz23bfailed=sz23bpresent=sz23babsent=0
        sz23cpassed=sz23cfailed=sz23cpresent=sz23cabsent=0
        sz231passed=sz231failed=sz231present=sz231absent=0
        sz33apassed=sz33afailed=sz33apresent=sz33aabsent=0
        tssecpassed=tssecfailed=tssecpresent=tssecabsent=0
        sz24apassed=sz24afailed=sz24apresent=sz24aabsent=0
        sz24bpassed=sz24bfailed=sz24bpresent=sz24babsent=0
        sz24cpassed=sz24cfailed=sz24cpresent=sz24cabsent=0
        sz241passed=sz241failed=sz241present=sz241absent=0
        sz34apassed=sz34afailed=sz34apresent=sz34aabsent=0
        tssedpassed=tssedfailed=tssedpresent=tssedabsent=0
        env4bpassed=env4bfailed=env4bpresent=env4babsent=0
        su25apassed=su25afailed=su25apresent=su25aabsent=0
        se25bpassed=se25bfailed=se25bpresent=se25babsent=0
        se25cpassed=se25cfailed=se25cpresent=se25cabsent=0
        se251passed=se251failed=se251present=se251absent=0
        se252passed=se252failed=se252present=se252absent=0
        sz45apassed=sz45afailed=sz45apresent=sz45aabsent=0
        vae5qpassed=vae5qfailed=vae5qpresent=vae5qabsent=0
        sz26apassed=sz26afailed=sz26apresent=sz26aabsent=0
        sz26bpassed=sz26bfailed=sz26bpresent=sz26babsent=0
        sz26cpassed=sz26cfailed=sz26cpresent=sz26cabsent=0
        sz261passed=sz261failed=sz261present=sz261absent=0
        su46bpassed=su46bfailed=su46bpresent=su46babsent=0
        sz26qpassed=sz26qfailed=sz26qpresent=sz26qabsent=0
        ces6qpassed=ces6qfailed=ces6qpresent=ces6qabsent=0

        la11a=cle1e=lk11a=lz11a=se21a=sm3aa=nlt1c=tlt1c=sm5aa=pz1sa=30
        la12a=cle2g=clk2t=lz12a=sm3ae=nlt2d=tlt2d=sm5ab=pz1sc=30
        sz23a=sz23b=sz23c=sz33a=30
        se25b=se25c=su25a=sz45a=vae5q=30
        se251=se252=sz241=se211=su22a=su221=sz231=24
        sz24a=sz24b=sz24c=sz34a=env4b=30
        tssec=tssed=20
        sz26a=sz26b=sz26c=su46b=30
        sz261=sz26q=24
        ces6q=1
        
        print("\n\n"+"CREATING"+" "+sheetname+" "+"SHEET"+"\n\n")
        sheetname = wb.create_sheet(sheetname)
        
        sheetname.append(['DATE OF BIRTH','NAME','REGISTER NUMBER','LA11A EXTERNAL' ,'LA11A INTERNAL' ,'LA11A TOTAL','CLE1E EXTERNAL' ,'CLE1E INTERNAL' ,'CLE1E TOTAL','LK11A EXTERNAL' ,'LK11A INTERNAL' ,'LK11A TOTAL','LZ11A EXTERNAL' ,'LZ11A INTERNAL' ,'LZ11A TOTAL','SE21A EXTERNAL' ,'SE21A INTERNAL' ,'SE21A TOTAL','SE211 EXTERNAL' ,'SE211 INTERNAL' ,'SE211 TOTAL','SM3AA EXTERNAL' ,'SM3AA INTERNAL' ,'SM3AA TOTAL','NLT1C EXTERNAL' ,'NLT1C INTERNAL' ,'NLT1C TOTAL','TLT1C EXTERNAL' ,'TLT1C INTERNAL' ,'TLT1C TOTAL','SM5AA EXTERNAL' ,'SM5AA INTERNAL' ,'SM5AA TOTAL','PZ1SA EXTERNAL' ,'PZ1SA INTERNAL' ,'PZ1SA TOTAL','LA12A EXTERNAL' ,'LA12A INTERNAL' ,'LA12A TOTAL','CLE2G EXTERNAL' ,'CLE2G INTERNAL' ,'CLE2G TOTAL','CLK2T EXTERNAL' ,'CLK2T INTERNAL' ,'CLK2T TOTAL','LZ12A EXTERNAL' ,'LZ12A INTERNAL' ,'LZ12A TOTAL','SU22A EXTERNAL' ,'SU22A INTERNAL' ,'SU22A TOTAL','SU221 EXTERNAL' ,'SU221 INTERNAL' ,'SU221 TOTAL','SM3AE EXTERNAL' ,'SM3AE INTERNAL' ,'SM3AE TOTAL','NLT2D EXTERNAL' ,'NLT2D INTERNAL' ,'NLT2D TOTAL','TLT2D EXTERNAL' ,'TLT2D INTERNAL' ,'TLT2D TOTAL','SM5AB EXTERNAL' ,'SM5AB INTERNAL' ,'SM5AB TOTAL','PZ1SC EXTERNAL' ,'PZ1SC INTERNAL' ,'PZ1SC TOTAL','SZ23A EXTERNAL' ,'SZ23A INTERNAL' ,'SZ23A TOTAL','SZ23B EXTERNAL' ,'SZ23B INTERNAL' ,'SZ23B TOTAL','SZ23C EXTERNAL' ,'SZ23C INTERNAL' ,'SZ23C TOTAL','SZ231 EXTERNAL' ,'SZ231 INTERNAL' ,'SZ231 TOTAL','SZ33A EXTERNAL' ,'SZ33A INTERNAL' ,'SZ33A TOTAL','TSSEC EXTERNAL' ,'TSSEC INTERNAL' ,'TSSEC TOTAL','SZ24A EXTERNAL' ,'SZ24A INTERNAL' ,'SZ24A TOTAL','SZ24B EXTERNAL' ,'SZ24B INTERNAL' ,'SZ24B TOTAL','SZ24C EXTERNAL' ,'SZ24C INTERNAL' ,'SZ24C TOTAL','SZ241 EXTERNAL' ,'SZ241 INTERNAL' ,'SZ241 TOTAL','SZ34A EXTERNAL' ,'SZ34A INTERNAL' ,'SZ34A TOTAL','TSSED EXTERNAL' ,'TSSED INTERNAL' ,'TSSED TOTAL','ENV4B EXTERNAL' ,'ENV4B INTERNAL' ,'ENV4B TOTAL','SU25A EXTERNAL' ,'SU25A INTERNAL' ,'SU25A TOTAL','SE25B EXTERNAL' ,'SE25B INTERNAL' ,'SE25B TOTAL','SE25C EXTERNAL' ,'SE25C INTERNAL' ,'SE25C TOTAL','SE251 EXTERNAL' ,'SE251 INTERNAL' ,'SE251 TOTAL','SE252 EXTERNAL' ,'SE252 INTERNAL' ,'SE252 TOTAL','SZ45A EXTERNAL' ,'SZ45A INTERNAL' ,'SZ45A TOTAL','VAE5Q EXTERNAL' ,'VAE5Q INTERNAL' ,'VAE5Q TOTAL', 'SZ26A EXTERNAL', 'SZ26A INTERNAL', 'SZ26A TOTAL', 'SZ26B EXTERNAL', 'SZ26B INTERNAL', 'SZ26B TOTAL', 'SZ26C EXTERNAL', 'SZ26C INTERNAL', 'SZ26C TOTAL', 'SZ261 EXTERNAL', 'SZ261 INTERNAL', 'SZ261 TOTAL', 'SU46B EXTERNAL', 'SU46B INTERNAL', 'SU46B TOTAL', 'SZ26Q EXTERNAL', 'SZ26Q INTERNAL', 'SZ26Q TOTAL', 'CES6Q EXTERNAL', 'CES6Q INTERNAL', 'CES6Q TOTAL'])

        for x,y in zip(rn,db):
                payload = {
                'regno': x,
                'pwd'  : y, 
                'button': 'Get Result'
        }
                url = 'https://results.unom.ac.in/results/ugresultpage.asp'
                data = urlencode(payload)
                data = data.encode('ascii')
                req = Request(url,data)
                req.add_header('User-Agent', 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Mobile Safari/537.36')
                res = urlopen(req)
                lxml = BeautifulSoup(res, 'html.parser')
                scraped=lxml.find_all("span", class_="TdsTxt")
                res_array=[res_array.text for res_array in scraped]
                subcode=['LA11A', 'CLE1E', 'LK11A', 'LZ11A', 'SE21A', 'SE211', 'SM3AA', 'NLT1C', 'TLT1C', 'SM5AA', 'PZ1SA', 'LA12A', 'CLE2G', 'CLK2T', 'LZ12A', 'SU22A', 'SU221', 'SM3AE', 'NLT2D', 'TLT2D', 'SM5AB', 'PZ1SC', 'SZ23A', 'SZ23B', 'SZ23C', 'SZ231', 'SZ33A', 'TSSEC', 'SZ24A', 'SZ24B', 'SZ24C', 'SZ241', 'SZ34A', 'TSSED', 'ENV4B', 'SU25A', 'SE25B', 'SE25C', 'SE251', 'SE252', 'SZ45A', 'VAE5Q', 'SZ26A', 'SZ26B', 'SZ26C', 'SZ261', 'SU46B', 'SZ26Q', 'CES6Q']
                sname=res_array[2].replace(" Name : ","")
                regino=res_array[3].replace("Register Number : ","")
                s_dob=res_array[4].replace("DOB :  ","")
                sheetname.append([s_dob,sname,regino])
                ro=ro+1
                count+=1 
                for length in range(0,len(subcode)):
                        spaced_subcode=" "+subcode[length]+" "
                        if spaced_subcode in res_array:
                                subcode_index=res_array.index(spaced_subcode)
                                subex=res_array[subcode_index+1]
                                subint=res_array[subcode_index+2]
                                subtot=res_array[subcode_index+3]
                                if subex=="000":
                                        subex=int(0)
                                elif subex=="AAA":      
                                        subex="AAA"
                                elif subex.isdigit():
                                        subex=int(subex.lstrip('0'))
                                if subint=="000":
                                        subint=int(0)
                                elif subint=="AAA":
                                        subint="AAA"
                                elif subint.isdigit():
                                        subint=int(subint.lstrip('0'))
                                if subtot=="000":
                                        subtot=int(0)
                                elif subtot=="AAA":
                                        subtot="AAA"
                                elif subtot=="***":
                                        subtot=int(subex)+int(subint)
                                elif subtot.isdigit():
                                        subtot=int(subtot.lstrip('0'))
                                if subcode[length] == subcode[0]:
                                        sheetname.cell(row=ro,column=4).value=subex 
                                        sheetname.cell(row=ro,column=5).value=subint
                                        sheetname.cell(row=ro,column=6).value=subtot
                                        if str(subex).isdigit():
                                                if int(subex)>=la11a:
                                                        la11apassed+=1
                                                        la11apresent+=1
                                                elif int(subex)<la11a:
                                                        la11apresent+=1
                                                        la11afailed+=1
                                        elif subex.isalpha():
                                                if subex=="AAA":
                                                        la11aabsent+=1
                                if subcode[length] == subcode[1]:
                                                        sheetname.cell(row=ro,column=7).value=subex 
                                                        sheetname.cell(row=ro,column=8).value=subint
                                                        sheetname.cell(row=ro,column=9).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=cle1e:
                                                                        cle1epassed+=1
                                                                        cle1epresent+=1
                                                                elif int(subex)<cle1e:
                                                                        cle1epresent+=1
                                                                        cle1efailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        cle1eabsent+=1
                                if subcode[length] == subcode[2]:
                                                        sheetname.cell(row=ro,column=10).value=subex
                                                        sheetname.cell(row=ro,column=11).value=subint
                                                        sheetname.cell(row=ro,column=12).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=lk11a:
                                                                        lk11apassed+=1
                                                                        lk11apresent+=1
                                                                elif int(subex)<lk11a:
                                                                        lk11apresent+=1
                                                                        lk11afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        lk11aabsent+=1
                                if subcode[length] == subcode[3]:
                                                        sheetname.cell(row=ro,column=13).value=subex
                                                        sheetname.cell(row=ro,column=14).value=subint
                                                        sheetname.cell(row=ro,column=15).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=lz11a:
                                                                        lz11apassed+=1
                                                                        lz11apresent+=1
                                                                elif int(subex)<lz11a:
                                                                        lz11apresent+=1
                                                                        lz11afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        lz11aabsent+=1
                                if subcode[length] == subcode[4]:
                                                        sheetname.cell(row=ro,column=16).value=subex
                                                        sheetname.cell(row=ro,column=17).value=subint
                                                        sheetname.cell(row=ro,column=18).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=se21a:
                                                                        se21apassed+=1
                                                                        se21apresent+=1
                                                                elif int(subex)<se21a:
                                                                        se21apresent+=1
                                                                        se21afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        se21aabsent+=1
                                if subcode[length] == subcode[5]:
                                                        sheetname.cell(row=ro,column=19).value=subex
                                                        sheetname.cell(row=ro,column=20).value=subint
                                                        sheetname.cell(row=ro,column=21).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=se211:
                                                                        se211passed+=1
                                                                        se211present+=1
                                                                elif int(subex)<se211:
                                                                        se211present+=1
                                                                        se211failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        se211absent+=1
                                if subcode[length] == subcode[6]:
                                                        sheetname.cell(row=ro,column=22).value=subex
                                                        sheetname.cell(row=ro,column=23).value=subint
                                                        sheetname.cell(row=ro,column=24).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sm3aa:
                                                                        sm3aapassed+=1
                                                                        sm3aapresent+=1
                                                                elif int(subex)<sm3aa:
                                                                        sm3aapresent+=1
                                                                        sm3aafailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sm3aaabsent+=1
                                if subcode[length] == subcode[7]:
                                                        sheetname.cell(row=ro,column=25).value=subex
                                                        sheetname.cell(row=ro,column=26).value=subint
                                                        sheetname.cell(row=ro,column=27).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=nlt1c:
                                                                        nlt1cpassed+=1
                                                                        nlt1cpresent+=1
                                                                elif int(subex)<nlt1c:
                                                                        nlt1cpresent+=1
                                                                        nlt1cfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        nlt1cabsent+=1
                                if subcode[length] == subcode[8]:
                                                        sheetname.cell(row=ro,column=28).value=subex
                                                        sheetname.cell(row=ro,column=29).value=subint
                                                        sheetname.cell(row=ro,column=30).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=tlt1c:
                                                                        tlt1cpassed+=1
                                                                        tlt1cpresent+=1
                                                                elif int(subex)<tlt1c:
                                                                        tlt1cpresent+=1
                                                                        tlt1cfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        tlt1cabsent+=1
                                if subcode[length] == subcode[9]:
                                                        sheetname.cell(row=ro,column=31).value=subex
                                                        sheetname.cell(row=ro,column=32).value=subint
                                                        sheetname.cell(row=ro,column=33).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sm5aa:
                                                                        sm5aapassed+=1
                                                                        sm5aapresent+=1
                                                                elif int(subex)<sm5aa:
                                                                        sm5aapresent+=1
                                                                        sm5aafailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sm5aaabsent+=1
                                if subcode[length] == subcode[10]:
                                                        sheetname.cell(row=ro,column=34).value=subex
                                                        sheetname.cell(row=ro,column=35).value=subint
                                                        sheetname.cell(row=ro,column=36).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=pz1sa:
                                                                        pz1sapassed+=1
                                                                        pz1sapresent+=1
                                                                elif int(subex)<pz1sa:
                                                                        pz1sapresent+=1
                                                                        pz1safailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        pz1saabsent+=1
                                if subcode[length] == subcode[11]:
                                                        sheetname.cell(row=ro,column=37).value=subex
                                                        sheetname.cell(row=ro,column=38).value=subint
                                                        sheetname.cell(row=ro,column=39).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=la12a:
                                                                        la12apassed+=1
                                                                        la12apresent+=1
                                                                elif int(subex)<la12a:
                                                                        la12apresent+=1
                                                                        la12afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        la12aabsent+=1
                                if subcode[length] == subcode[12]:
                                                        sheetname.cell(row=ro,column=40).value=subex
                                                        sheetname.cell(row=ro,column=41).value=subint
                                                        sheetname.cell(row=ro,column=42).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=cle2g:
                                                                        cle2gpassed+=1
                                                                        cle2gpresent+=1
                                                                elif int(subex)<cle2g:
                                                                        cle2gpresent+=1
                                                                        cle2gfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        cle2gabsent+=1
                                if subcode[length] == subcode[13]:
                                                        sheetname.cell(row=ro,column=43).value=subex
                                                        sheetname.cell(row=ro,column=44).value=subint
                                                        sheetname.cell(row=ro,column=45).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=clk2t:
                                                                        clk2tpassed+=1
                                                                        clk2tpresent+=1
                                                                elif int(subex)<clk2t:
                                                                        clk2tpresent+=1
                                                                        clk2tfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        clk2tabsent+=1
                                if subcode[length] == subcode[14]:
                                                        sheetname.cell(row=ro,column=46).value=subex
                                                        sheetname.cell(row=ro,column=47).value=subint
                                                        sheetname.cell(row=ro,column=48).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=lz12a:
                                                                        lz12apassed+=1
                                                                        lz12apresent+=1
                                                                elif int(subex)<lz12a:
                                                                        lz12apresent+=1
                                                                        lz12afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        lz12aabsent+=1
                                if subcode[length] == subcode[15]:
                                                        sheetname.cell(row=ro,column=49).value=subex
                                                        sheetname.cell(row=ro,column=50).value=subint
                                                        sheetname.cell(row=ro,column=51).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=su22a:
                                                                        su22apassed+=1
                                                                        su22apresent+=1
                                                                elif int(subex)<su22a:
                                                                        su22apresent+=1
                                                                        su22afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        su22aabsent+=1
                                if subcode[length] == subcode[16]:
                                                        sheetname.cell(row=ro,column=52).value=subex
                                                        sheetname.cell(row=ro,column=53).value=subint
                                                        sheetname.cell(row=ro,column=54).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=su221:
                                                                        su221passed+=1
                                                                        su221present+=1
                                                                elif int(subex)<su221:
                                                                        su221present+=1
                                                                        su221failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        su221absent+=1
                                if subcode[length] == subcode[17]:
                                                        sheetname.cell(row=ro,column=55).value=subex
                                                        sheetname.cell(row=ro,column=56).value=subint
                                                        sheetname.cell(row=ro,column=57).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sm3ae:
                                                                        sm3aepassed+=1
                                                                        sm3aepresent+=1
                                                                elif int(subex)<sm3ae:
                                                                        sm3aepresent+=1
                                                                        sm3aefailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sm3aeabsent+=1
                                if subcode[length] == subcode[18]:
                                                        sheetname.cell(row=ro,column=58).value=subex
                                                        sheetname.cell(row=ro,column=59).value=subint
                                                        sheetname.cell(row=ro,column=60).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=nlt2d:
                                                                        nlt2dpassed+=1
                                                                        nlt2dpresent+=1
                                                                elif int(subex)<nlt2d:
                                                                        nlt2dpresent+=1
                                                                        nlt2dfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        nlt2dabsent+=1
                                if subcode[length] == subcode[19]:
                                                        sheetname.cell(row=ro,column=61).value=subex
                                                        sheetname.cell(row=ro,column=62).value=subint
                                                        sheetname.cell(row=ro,column=63).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=tlt2d:
                                                                        tlt2dpassed+=1
                                                                        tlt2dpresent+=1
                                                                elif int(subex)<tlt2d:
                                                                        tlt2dpresent+=1
                                                                        tlt2dfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        tlt2dabsent+=1
                                if subcode[length] == subcode[20]:
                                                        sheetname.cell(row=ro,column=64).value=subex
                                                        sheetname.cell(row=ro,column=65).value=subint
                                                        sheetname.cell(row=ro,column=66).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sm5ab:
                                                                        sm5abpassed+=1
                                                                        sm5abpresent+=1
                                                                elif int(subex)<sm5ab:
                                                                        sm5abpresent+=1
                                                                        sm5abfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sm5ababsent+=1
                                if subcode[length] == subcode[21]:
                                                        sheetname.cell(row=ro,column=67).value=subex
                                                        sheetname.cell(row=ro,column=68).value=subint
                                                        sheetname.cell(row=ro,column=69).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=pz1sc:
                                                                        pz1scpassed+=1
                                                                        pz1scpresent+=1
                                                                elif int(subex)<pz1sc:
                                                                        pz1scpresent+=1
                                                                        pz1scfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        pz1scabsent+=1
                                if subcode[length] == subcode[22]:
                                                        sheetname.cell(row=ro,column=70).value=subex
                                                        sheetname.cell(row=ro,column=71).value=subint
                                                        sheetname.cell(row=ro,column=72).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz23a:
                                                                        sz23apassed+=1
                                                                        sz23apresent+=1
                                                                elif int(subex)<sz23a:
                                                                        sz23apresent+=1
                                                                        sz23afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz23aabsent+=1
                                if subcode[length] == subcode[23]:
                                                        sheetname.cell(row=ro,column=73).value=subex
                                                        sheetname.cell(row=ro,column=74).value=subint
                                                        sheetname.cell(row=ro,column=75).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz23b:
                                                                        sz23bpassed+=1
                                                                        sz23bpresent+=1
                                                                elif int(subex)<sz23b:
                                                                        sz23bpresent+=1
                                                                        sz23bfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz23babsent+=1
                                if subcode[length] == subcode[24]:
                                                        sheetname.cell(row=ro,column=76).value=subex
                                                        sheetname.cell(row=ro,column=77).value=subint
                                                        sheetname.cell(row=ro,column=78).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz23c:
                                                                        sz23cpassed+=1
                                                                        sz23cpresent+=1
                                                                elif int(subex)<sz23c:
                                                                        sz23cpresent+=1
                                                                        sz23cfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz23cabsent+=1
                                if subcode[length] == subcode[25]:
                                                        sheetname.cell(row=ro,column=79).value=subex
                                                        sheetname.cell(row=ro,column=80).value=subint
                                                        sheetname.cell(row=ro,column=81).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz231:
                                                                        sz231passed+=1
                                                                        sz231present+=1
                                                                elif int(subex)<sz231:
                                                                        sz231present+=1
                                                                        sz231failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz231absent+=1
                                if subcode[length] == subcode[26]:
                                                        sheetname.cell(row=ro,column=82).value=subex
                                                        sheetname.cell(row=ro,column=83).value=subint
                                                        sheetname.cell(row=ro,column=84).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz33a:
                                                                        sz33apassed+=1
                                                                        sz33apresent+=1
                                                                elif int(subex)<sz33a:
                                                                        sz33apresent+=1
                                                                        sz33afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz33aabsent+=1
                                if subcode[length] == subcode[27]:
                                                        sheetname.cell(row=ro,column=85).value=subex
                                                        sheetname.cell(row=ro,column=86).value=subint
                                                        sheetname.cell(row=ro,column=87).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=tssec:
                                                                        tssecpassed+=1
                                                                        tssecpresent+=1
                                                                elif int(subex)<tssec:
                                                                        tssecpresent+=1
                                                                        tssecfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        tssecabsent+=1
                                if subcode[length] == subcode[28]:
                                                        sheetname.cell(row=ro,column=88).value=subex
                                                        sheetname.cell(row=ro,column=89).value=subint
                                                        sheetname.cell(row=ro,column=90).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz24a:
                                                                        sz24apassed+=1
                                                                        sz24apresent+=1
                                                                elif int(subex)<sz24a:
                                                                        sz24apresent+=1
                                                                        sz24afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz24aabsent+=1
                                if subcode[length] == subcode[29]:
                                                        sheetname.cell(row=ro,column=91).value=subex
                                                        sheetname.cell(row=ro,column=92).value=subint
                                                        sheetname.cell(row=ro,column=93).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz24b:
                                                                        sz24bpassed+=1
                                                                        sz24bpresent+=1
                                                                elif int(subex)<sz24b:
                                                                        sz24bpresent+=1
                                                                        sz24bfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz24babsent+=1
                                if subcode[length] == subcode[30]:
                                                        sheetname.cell(row=ro,column=94).value=subex
                                                        sheetname.cell(row=ro,column=95).value=subint
                                                        sheetname.cell(row=ro,column=96).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz24c:
                                                                        sz24cpassed+=1
                                                                        sz24cpresent+=1
                                                                elif int(subex)<sz24c:
                                                                        sz24cpresent+=1
                                                                        sz24cfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz24cabsent+=1
                                if subcode[length] == subcode[31]:
                                                        sheetname.cell(row=ro,column=97).value=subex
                                                        sheetname.cell(row=ro,column=98).value=subint
                                                        sheetname.cell(row=ro,column=99).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz241:
                                                                        sz241passed+=1
                                                                        sz241present+=1
                                                                elif int(subex)<sz241:
                                                                        sz241present+=1
                                                                        sz241failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz241absent+=1
                                if subcode[length] == subcode[32]:
                                                        sheetname.cell(row=ro,column=100).value=subex
                                                        sheetname.cell(row=ro,column=101).value=subint
                                                        sheetname.cell(row=ro,column=102).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz34a:
                                                                        sz34apassed+=1
                                                                        sz34apresent+=1
                                                                elif int(subex)<sz34a:
                                                                        sz34apresent+=1
                                                                        sz34afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz34aabsent+=1
                                if subcode[length] == subcode[33]:
                                                        sheetname.cell(row=ro,column=103).value=subex
                                                        sheetname.cell(row=ro,column=104).value=subint
                                                        sheetname.cell(row=ro,column=105).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=tssed:
                                                                        tssedpassed+=1
                                                                        tssedpresent+=1
                                                                elif int(subex)<tssed:
                                                                        tssedpresent+=1
                                                                        tssedfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        tssedabsent+=1
                                if subcode[length] == subcode[34]:
                                                        sheetname.cell(row=ro,column=106).value=subex
                                                        sheetname.cell(row=ro,column=107).value=subint
                                                        sheetname.cell(row=ro,column=108).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=env4b:
                                                                        env4bpassed+=1
                                                                        env4bpresent+=1
                                                                elif int(subex)<env4b:
                                                                        env4bpresent+=1
                                                                        env4bfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        env4babsent+=1
                                if subcode[length] == subcode[35]:
                                                        sheetname.cell(row=ro,column=109).value=subex
                                                        sheetname.cell(row=ro,column=110).value=subint
                                                        sheetname.cell(row=ro,column=111).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=su25a:
                                                                        su25apassed+=1
                                                                        su25apresent+=1
                                                                elif int(subex)<su25a:
                                                                        su25apresent+=1
                                                                        su25afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        su25aabsent+=1
                                if subcode[length] == subcode[36]:
                                                        sheetname.cell(row=ro,column=112).value=subex
                                                        sheetname.cell(row=ro,column=113).value=subint
                                                        sheetname.cell(row=ro,column=114).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=se25b:
                                                                        se25bpassed+=1
                                                                        se25bpresent+=1
                                                                elif int(subex)<se25b:
                                                                        se25bpresent+=1
                                                                        se25bfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        se25babsent+=1
                                if subcode[length] == subcode[37]:
                                                        sheetname.cell(row=ro,column=115).value=subex
                                                        sheetname.cell(row=ro,column=116).value=subint
                                                        sheetname.cell(row=ro,column=117).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=se25c:
                                                                        se25cpassed+=1
                                                                        se25cpresent+=1
                                                                elif int(subex)<se25c:
                                                                        se25cpresent+=1
                                                                        se25cfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        se25cabsent+=1
                                if subcode[length] == subcode[38]:
                                                        sheetname.cell(row=ro,column=118).value=subex
                                                        sheetname.cell(row=ro,column=119).value=subint
                                                        sheetname.cell(row=ro,column=120).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=se251:
                                                                        se251passed+=1
                                                                        se251present+=1
                                                                elif int(subex)<se251:
                                                                        se251present+=1
                                                                        se251failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        se251absent+=1
                                if subcode[length] == subcode[39]:
                                                        sheetname.cell(row=ro,column=121).value=subex
                                                        sheetname.cell(row=ro,column=122).value=subint
                                                        sheetname.cell(row=ro,column=123).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=se252:
                                                                        se252passed+=1
                                                                        se252present+=1
                                                                elif int(subex)<se252:
                                                                        se252present+=1
                                                                        se252failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        se252absent+=1
                                if subcode[length] == subcode[40]:
                                                        sheetname.cell(row=ro,column=124).value=subex
                                                        sheetname.cell(row=ro,column=125).value=subint
                                                        sheetname.cell(row=ro,column=126).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz45a:
                                                                        sz45apassed+=1
                                                                        sz45apresent+=1
                                                                elif int(subex)<sz45a:
                                                                        sz45apresent+=1
                                                                        sz45afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz45aabsent+=1
                                if subcode[length] == subcode[41]:
                                                        sheetname.cell(row=ro,column=127).value=subex
                                                        sheetname.cell(row=ro,column=128).value=subint
                                                        sheetname.cell(row=ro,column=129).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=vae5q:
                                                                        vae5qpassed+=1
                                                                        vae5qpresent+=1
                                                                elif int(subex)<vae5q:
                                                                        vae5qpresent+=1
                                                                        vae5qfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        vae5qabsent+=1
                                if subcode[length] == subcode[42]:
                                                        sheetname.cell(row=ro,column=130).value=subex
                                                        sheetname.cell(row=ro,column=131).value=subint
                                                        sheetname.cell(row=ro,column=132).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz26a:
                                                                        sz26apassed+=1
                                                                        sz26apresent+=1
                                                                elif int(subex)<sz26a:
                                                                       sz26apresent+=1
                                                                       sz26afailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz26aabsent+=1
                                if subcode[length] == subcode[43]:
                                                        sheetname.cell(row=ro,column=133).value=subex
                                                        sheetname.cell(row=ro,column=134).value=subint
                                                        sheetname.cell(row=ro,column=135).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz26b:
                                                                        sz26bpassed+=1
                                                                        sz26bpresent+=1
                                                                elif int(subex)<sz26b:
                                                                       sz26bpresent+=1
                                                                       sz26bfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz26babsent+=1
                                if subcode[length] == subcode[44]:
                                                        sheetname.cell(row=ro,column=136).value=subex
                                                        sheetname.cell(row=ro,column=137).value=subint
                                                        sheetname.cell(row=ro,column=138).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz26c:
                                                                        sz26cpassed+=1
                                                                        sz26cpresent+=1
                                                                elif int(subex)<sz26c:
                                                                       sz26cpresent+=1
                                                                       sz26cfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz26cabsent+=1
                                if subcode[length] == subcode[45]:
                                                        sheetname.cell(row=ro,column=139).value=subex
                                                        sheetname.cell(row=ro,column=140).value=subint
                                                        sheetname.cell(row=ro,column=141).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz261:
                                                                        sz261passed+=1
                                                                        sz261present+=1
                                                                elif int(subex)<sz261:
                                                                       sz261present+=1
                                                                       sz261failed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz261absent+=1
                                if subcode[length] == subcode[46]:
                                                        sheetname.cell(row=ro,column=142).value=subex
                                                        sheetname.cell(row=ro,column=143).value=subint
                                                        sheetname.cell(row=ro,column=144).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=su46b:
                                                                        su46bpassed+=1
                                                                        su46bpresent+=1
                                                                elif int(subex)<su46b:
                                                                       su46bpresent+=1
                                                                       su46bfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        su46babsent+=1
                                if subcode[length] == subcode[47]:
                                                        sheetname.cell(row=ro,column=145).value=subex
                                                        sheetname.cell(row=ro,column=146).value=subint
                                                        sheetname.cell(row=ro,column=147).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=sz26q:
                                                                        sz26qpassed+=1
                                                                        sz26qpresent+=1
                                                                elif int(subex)<sz26q:
                                                                       sz26qpresent+=1
                                                                       sz26qfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        sz26qabsent+=1
                                if subcode[length] == subcode[48]:
                                                        sheetname.cell(row=ro,column=148).value=subex
                                                        sheetname.cell(row=ro,column=149).value=subint
                                                        sheetname.cell(row=ro,column=150).value=subtot
                                                        if str(subex).isdigit():
                                                                if int(subex)>=ces6q:
                                                                        ces6qpassed+=1
                                                                        ces6qpresent+=1
                                                                elif int(subex)<ces6q:
                                                                       ces6qpresent+=1
                                                                       ces6qfailed+=1
                                                        elif subex.isalpha():
                                                                if subex=="AAA":
                                                                        ces6qabsent+=1                                                                                                                                                                                                                                                                                                                                                                                                                                                                       

                print(str(count)+"."+" "+regino+" "+sname+" "+"appended")
        
        sheetname.cell(row=60,column=3).value="no of present"
        sheetname.cell(row=61,column=3).value="no of absent"
        sheetname.cell(row=62,column=3).value="no of passed"
        sheetname.cell(row=63,column=3).value="no of failed"
        sheetname.cell(row=64,column=3).value="pass %"

        la11apassper=(la11apassed/la11apresent)*100 if la11apassed and la11apresent !=0 else 0

        sheetname.cell(row=60,column=6).value=la11apresent
        sheetname.cell(row=61,column=6).value=la11aabsent
        sheetname.cell(row=62,column=6).value=la11apassed
        sheetname.cell(row=63,column=6).value=la11afailed
        sheetname.cell(row=64,column=6).value=la11apassper

        cle1epassper=(cle1epassed/cle1epresent)*100 if cle1epassed and cle1epresent !=0 else 0

        sheetname.cell(row=60,column=9).value=cle1epresent
        sheetname.cell(row=61,column=9).value=cle1eabsent
        sheetname.cell(row=62,column=9).value=cle1epassed
        sheetname.cell(row=63,column=9).value=cle1efailed
        sheetname.cell(row=64,column=9).value=cle1epassper

        lk11apassper=(lk11apassed/lk11apresent)*100 if lk11apassed and lk11apresent !=0 else 0

        sheetname.cell(row=60,column=12).value=lk11apresent
        sheetname.cell(row=61,column=12).value=lk11aabsent
        sheetname.cell(row=62,column=12).value=lk11apassed
        sheetname.cell(row=63,column=12).value=lk11afailed
        sheetname.cell(row=64,column=12).value=lk11apassper

        lz11apassper=(lz11apassed/lz11apresent)*100 if lz11apassed and lz11apresent !=0 else 0

        sheetname.cell(row=60,column=15).value=lz11apresent
        sheetname.cell(row=61,column=15).value=lz11aabsent
        sheetname.cell(row=62,column=15).value=lz11apassed
        sheetname.cell(row=63,column=15).value=lz11afailed
        sheetname.cell(row=64,column=15).value=lz11apassper

        se21apassper=(se21apassed/se21apresent)*100 if se21apassed and se21apresent !=0 else 0

        sheetname.cell(row=60,column=18).value=se21apresent
        sheetname.cell(row=61,column=18).value=se21aabsent
        sheetname.cell(row=62,column=18).value=se21apassed
        sheetname.cell(row=63,column=18).value=se21afailed
        sheetname.cell(row=64,column=18).value=se21apassper

        se211passper=(se211passed/se211present)*100 if se211passed and se211present !=0 else 0

        sheetname.cell(row=60,column=21).value=se211present
        sheetname.cell(row=61,column=21).value=se211absent
        sheetname.cell(row=62,column=21).value=se211passed
        sheetname.cell(row=63,column=21).value=se211failed
        sheetname.cell(row=64,column=21).value=se211passper

        sm3aapassper=(sm3aapassed/sm3aapresent)*100 if sm3aapassed and sm3aapresent !=0 else 0

        sheetname.cell(row=60,column=24).value=sm3aapresent
        sheetname.cell(row=61,column=24).value=sm3aaabsent
        sheetname.cell(row=62,column=24).value=sm3aapassed
        sheetname.cell(row=63,column=24).value=sm3aafailed
        sheetname.cell(row=64,column=24).value=sm3aapassper

        nlt1cpassper=(nlt1cpassed/nlt1cpresent)*100 if nlt1cpassed and nlt1cpresent !=0 else 0

        sheetname.cell(row=60,column=27).value=nlt1cpresent
        sheetname.cell(row=61,column=27).value=nlt1cabsent
        sheetname.cell(row=62,column=27).value=nlt1cpassed
        sheetname.cell(row=63,column=27).value=nlt1cfailed
        sheetname.cell(row=64,column=27).value=nlt1cpassper

        tlt1cpassper=(tlt1cpassed/tlt1cpresent)*100 if tlt1cpassed and tlt1cpresent !=0 else 0

        sheetname.cell(row=60,column=30).value=tlt1cpresent
        sheetname.cell(row=61,column=30).value=tlt1cabsent
        sheetname.cell(row=62,column=30).value=tlt1cpassed
        sheetname.cell(row=63,column=30).value=tlt1cfailed
        sheetname.cell(row=64,column=30).value=tlt1cpassper

        sm5aapassper=(sm5aapassed/sm5aapresent)*100 if sm5aapassed and sm5aapresent !=0 else 0

        sheetname.cell(row=60,column=33).value=sm5aapresent
        sheetname.cell(row=61,column=33).value=sm5aaabsent
        sheetname.cell(row=62,column=33).value=sm5aapassed
        sheetname.cell(row=63,column=33).value=sm5aafailed
        sheetname.cell(row=64,column=33).value=sm5aapassper

        pz1sapassper=(pz1sapassed/pz1sapresent)*100 if pz1sapassed and pz1sapresent !=0 else 0

        sheetname.cell(row=60,column=36).value=pz1sapresent
        sheetname.cell(row=61,column=36).value=pz1saabsent
        sheetname.cell(row=62,column=36).value=pz1sapassed
        sheetname.cell(row=63,column=36).value=pz1safailed
        sheetname.cell(row=64,column=36).value=pz1sapassper

        la12apassper=(la12apassed/la12apresent)*100 if la12apassed and la12apresent !=0 else 0

        sheetname.cell(row=60,column=39).value=la12apresent
        sheetname.cell(row=61,column=39).value=la12aabsent
        sheetname.cell(row=62,column=39).value=la12apassed
        sheetname.cell(row=63,column=39).value=la12afailed
        sheetname.cell(row=64,column=39).value=la12apassper

        cle2gpassper=(cle2gpassed/cle2gpresent)*100 if cle2gpassed and cle2gpresent !=0 else 0

        sheetname.cell(row=60,column=42).value=cle2gpresent
        sheetname.cell(row=61,column=42).value=cle2gabsent
        sheetname.cell(row=62,column=42).value=cle2gpassed
        sheetname.cell(row=63,column=42).value=cle2gfailed
        sheetname.cell(row=64,column=42).value=cle2gpassper

        clk2tpassper=(clk2tpassed/clk2tpresent)*100 if clk2tpassed and clk2tpresent !=0 else 0

        sheetname.cell(row=60,column=45).value=clk2tpresent
        sheetname.cell(row=61,column=45).value=clk2tabsent
        sheetname.cell(row=62,column=45).value=clk2tpassed
        sheetname.cell(row=63,column=45).value=clk2tfailed
        sheetname.cell(row=64,column=45).value=clk2tpassper

        lz12apassper=(lz12apassed/lz12apresent)*100 if lz12apassed and lz12apresent !=0 else 0

        sheetname.cell(row=60,column=48).value=lz12apresent
        sheetname.cell(row=61,column=48).value=lz12aabsent
        sheetname.cell(row=62,column=48).value=lz12apassed
        sheetname.cell(row=63,column=48).value=lz12afailed
        sheetname.cell(row=64,column=48).value=lz12apassper

        su22apassper=(su22apassed/su22apresent)*100 if su22apassed and su22apresent !=0 else 0

        sheetname.cell(row=60,column=51).value=su22apresent
        sheetname.cell(row=61,column=51).value=su22aabsent
        sheetname.cell(row=62,column=51).value=su22apassed
        sheetname.cell(row=63,column=51).value=su22afailed
        sheetname.cell(row=64,column=51).value=su22apassper

        su221passper=(su221passed/su221present)*100 if su221passed and su221present !=0 else 0

        sheetname.cell(row=60,column=54).value=su221present
        sheetname.cell(row=61,column=54).value=su221absent
        sheetname.cell(row=62,column=54).value=su221passed
        sheetname.cell(row=63,column=54).value=su221failed
        sheetname.cell(row=64,column=54).value=su221passper

        sm3aepassper=(sm3aepassed/sm3aepresent)*100 if sm3aepassed and sm3aepresent !=0 else 0

        sheetname.cell(row=60,column=57).value=sm3aepresent
        sheetname.cell(row=61,column=57).value=sm3aeabsent
        sheetname.cell(row=62,column=57).value=sm3aepassed
        sheetname.cell(row=63,column=57).value=sm3aefailed
        sheetname.cell(row=64,column=57).value=sm3aepassper

        nlt2dpassper=(nlt2dpassed/nlt2dpresent)*100 if nlt2dpassed and nlt2dpresent !=0 else 0

        sheetname.cell(row=60,column=60).value=nlt2dpresent
        sheetname.cell(row=61,column=60).value=nlt2dabsent
        sheetname.cell(row=62,column=60).value=nlt2dpassed
        sheetname.cell(row=63,column=60).value=nlt2dfailed
        sheetname.cell(row=64,column=60).value=nlt2dpassper

        tlt2dpassper=(tlt2dpassed/tlt2dpresent)*100 if tlt2dpassed and tlt2dpresent !=0 else 0

        sheetname.cell(row=60,column=63).value=tlt2dpresent
        sheetname.cell(row=61,column=63).value=tlt2dabsent
        sheetname.cell(row=62,column=63).value=tlt2dpassed
        sheetname.cell(row=63,column=63).value=tlt2dfailed
        sheetname.cell(row=64,column=63).value=tlt2dpassper

        sm5abpassper=(sm5abpassed/sm5abpresent)*100 if sm5abpassed and sm5abpresent !=0 else 0

        sheetname.cell(row=60,column=66).value=sm5abpresent
        sheetname.cell(row=61,column=66).value=sm5ababsent
        sheetname.cell(row=62,column=66).value=sm5abpassed
        sheetname.cell(row=63,column=66).value=sm5abfailed
        sheetname.cell(row=64,column=66).value=sm5abpassper

        pz1scpassper=(pz1scpassed/pz1scpresent)*100 if pz1scpassed and pz1scpresent !=0 else 0

        sheetname.cell(row=60,column=69).value=pz1scpresent
        sheetname.cell(row=61,column=69).value=pz1scabsent
        sheetname.cell(row=62,column=69).value=pz1scpassed
        sheetname.cell(row=63,column=69).value=pz1scfailed
        sheetname.cell(row=64,column=69).value=pz1scpassper

        sz23apassper=(sz23apassed/sz23apresent)*100 if sz23apassed and sz23apresent !=0 else 0

        sheetname.cell(row=60,column=72).value=sz23apresent
        sheetname.cell(row=61,column=72).value=sz23aabsent
        sheetname.cell(row=62,column=72).value=sz23apassed
        sheetname.cell(row=63,column=72).value=sz23afailed
        sheetname.cell(row=64,column=72).value=sz23apassper

        sz23bpassper=(sz23bpassed/sz23bpresent)*100 if sz23bpassed and sz23bpresent !=0 else 0

        sheetname.cell(row=60,column=75).value=sz23bpresent
        sheetname.cell(row=61,column=75).value=sz23babsent
        sheetname.cell(row=62,column=75).value=sz23bpassed
        sheetname.cell(row=63,column=75).value=sz23bfailed
        sheetname.cell(row=64,column=75).value=sz23bpassper

        sz23cpassper=(sz23cpassed/sz23cpresent)*100 if sz23cpassed and sz23cpresent !=0 else 0

        sheetname.cell(row=60,column=78).value=sz23cpresent
        sheetname.cell(row=61,column=78).value=sz23cabsent
        sheetname.cell(row=62,column=78).value=sz23cpassed
        sheetname.cell(row=63,column=78).value=sz23cfailed
        sheetname.cell(row=64,column=78).value=sz23cpassper

        sz231passper=(sz231passed/sz231present)*100 if sz231passed and sz231present !=0 else 0

        sheetname.cell(row=60,column=81).value=sz231present
        sheetname.cell(row=61,column=81).value=sz231absent
        sheetname.cell(row=62,column=81).value=sz231passed
        sheetname.cell(row=63,column=81).value=sz231failed
        sheetname.cell(row=64,column=81).value=sz231passper

        sz33apassper=(sz33apassed/sz33apresent)*100 if sz33apassed and sz33apresent !=0 else 0

        sheetname.cell(row=60,column=84).value=sz33apresent
        sheetname.cell(row=61,column=84).value=sz33aabsent
        sheetname.cell(row=62,column=84).value=sz33apassed
        sheetname.cell(row=63,column=84).value=sz33afailed
        sheetname.cell(row=64,column=84).value=sz33apassper

        tssecpassper=(tssecpassed/tssecpresent)*100 if tssecpassed and tssecpresent !=0 else 0

        sheetname.cell(row=60,column=87).value=tssecpresent
        sheetname.cell(row=61,column=87).value=tssecabsent
        sheetname.cell(row=62,column=87).value=tssecpassed
        sheetname.cell(row=63,column=87).value=tssecfailed
        sheetname.cell(row=64,column=87).value=tssecpassper

        sz24apassper=(sz24apassed/sz24apresent)*100 if sz24apassed and sz24apresent !=0 else 0

        sheetname.cell(row=60,column=90).value=sz24apresent
        sheetname.cell(row=61,column=90).value=sz24aabsent
        sheetname.cell(row=62,column=90).value=sz24apassed
        sheetname.cell(row=63,column=90).value=sz24afailed
        sheetname.cell(row=64,column=90).value=sz24apassper

        sz24bpassper=(sz24bpassed/sz24bpresent)*100 if sz24bpassed and sz24bpresent !=0 else 0

        sheetname.cell(row=60,column=93).value=sz24bpresent
        sheetname.cell(row=61,column=93).value=sz24babsent
        sheetname.cell(row=62,column=93).value=sz24bpassed
        sheetname.cell(row=63,column=93).value=sz24bfailed
        sheetname.cell(row=64,column=93).value=sz24bpassper

        sz24cpassper=(sz24cpassed/sz24cpresent)*100 if sz24cpassed and sz24cpresent !=0 else 0

        sheetname.cell(row=60,column=96).value=sz24cpresent
        sheetname.cell(row=61,column=96).value=sz24cabsent
        sheetname.cell(row=62,column=96).value=sz24cpassed
        sheetname.cell(row=63,column=96).value=sz24cfailed
        sheetname.cell(row=64,column=96).value=sz24cpassper

        sz241passper=(sz241passed/sz241present)*100 if sz241passed and sz241present !=0 else 0

        sheetname.cell(row=60,column=99).value=sz241present
        sheetname.cell(row=61,column=99).value=sz241absent
        sheetname.cell(row=62,column=99).value=sz241passed
        sheetname.cell(row=63,column=99).value=sz241failed
        sheetname.cell(row=64,column=99).value=sz241passper

        sz34apassper=(sz34apassed/sz34apresent)*100 if sz34apassed and sz34apresent !=0 else 0

        sheetname.cell(row=60,column=102).value=sz34apresent
        sheetname.cell(row=61,column=102).value=sz34aabsent
        sheetname.cell(row=62,column=102).value=sz34apassed
        sheetname.cell(row=63,column=102).value=sz34afailed
        sheetname.cell(row=64,column=102).value=sz34apassper

        tssedpassper=(tssedpassed/tssedpresent)*100 if tssedpassed and tssedpresent !=0 else 0

        sheetname.cell(row=60,column=105).value=tssedpresent
        sheetname.cell(row=61,column=105).value=tssedabsent
        sheetname.cell(row=62,column=105).value=tssedpassed
        sheetname.cell(row=63,column=105).value=tssedfailed
        sheetname.cell(row=64,column=105).value=tssedpassper

        env4bpassper=(env4bpassed/env4bpresent)*100 if env4bpassed and env4bpresent !=0 else 0

        sheetname.cell(row=60,column=108).value=env4bpresent
        sheetname.cell(row=61,column=108).value=env4babsent
        sheetname.cell(row=62,column=108).value=env4bpassed
        sheetname.cell(row=63,column=108).value=env4bfailed
        sheetname.cell(row=64,column=108).value=env4bpassper

        su25apassper=(su25apassed/su25apresent)*100 if su25apassed and su25apresent !=0 else 0

        sheetname.cell(row=60,column=111).value=su25apresent
        sheetname.cell(row=61,column=111).value=su25aabsent
        sheetname.cell(row=62,column=111).value=su25apassed
        sheetname.cell(row=63,column=111).value=su25afailed
        sheetname.cell(row=64,column=111).value=su25apassper

        se25bpassper=(se25bpassed/se25bpresent)*100 if se25bpassed and se25bpresent !=0 else 0

        sheetname.cell(row=60,column=114).value=se25bpresent
        sheetname.cell(row=61,column=114).value=se25babsent
        sheetname.cell(row=62,column=114).value=se25bpassed
        sheetname.cell(row=63,column=114).value=se25bfailed
        sheetname.cell(row=64,column=114).value=se25bpassper

        se25cpassper=(se25cpassed/se25cpresent)*100 if se25cpassed and se25cpresent !=0 else 0

        sheetname.cell(row=60,column=117).value=se25cpresent
        sheetname.cell(row=61,column=117).value=se25cabsent
        sheetname.cell(row=62,column=117).value=se25cpassed
        sheetname.cell(row=63,column=117).value=se25cfailed
        sheetname.cell(row=64,column=117).value=se25cpassper

        se251passper=(se251passed/se251present)*100 if se251passed and se251present !=0 else 0

        sheetname.cell(row=60,column=120).value=se251present
        sheetname.cell(row=61,column=120).value=se251absent
        sheetname.cell(row=62,column=120).value=se251passed
        sheetname.cell(row=63,column=120).value=se251failed
        sheetname.cell(row=64,column=120).value=se251passper

        se252passper=(se252passed/se252present)*100 if se252passed and se252present !=0 else 0

        sheetname.cell(row=60,column=123).value=se252present
        sheetname.cell(row=61,column=123).value=se252absent
        sheetname.cell(row=62,column=123).value=se252passed
        sheetname.cell(row=63,column=123).value=se252failed
        sheetname.cell(row=64,column=123).value=se252passper

        sz45apassper=(sz45apassed/sz45apresent)*100 if sz45apassed and sz45apresent !=0 else 0

        sheetname.cell(row=60,column=126).value=sz45apresent
        sheetname.cell(row=61,column=126).value=sz45aabsent
        sheetname.cell(row=62,column=126).value=sz45apassed
        sheetname.cell(row=63,column=126).value=sz45afailed
        sheetname.cell(row=64,column=126).value=sz45apassper

        vae5qpassper=(vae5qpassed/vae5qpresent)*100 if vae5qpassed and vae5qpresent !=0 else 0

        sheetname.cell(row=60,column=129).value=vae5qpresent
        sheetname.cell(row=61,column=129).value=vae5qabsent
        sheetname.cell(row=62,column=129).value=vae5qpassed 
        sheetname.cell(row=63,column=129).value=vae5qfailed
        sheetname.cell(row=64,column=129).value=vae5qpassper

        sz26apassper=(sz26apassed/sz26apresent)*100 if sz26apassed and sz26apresent !=0 else 0

        sheetname.cell(row=60,column=132).value=sz26apresent
        sheetname.cell(row=61,column=132).value=sz26aabsent
        sheetname.cell(row=62,column=132).value=sz26apassed
        sheetname.cell(row=63,column=132).value=sz26afailed
        sheetname.cell(row=64,column=132).value=sz26apassper

        sz26bpassper=(sz26bpassed/sz26bpresent)*100 if sz26bpassed and sz26bpresent !=0 else 0

        sheetname.cell(row=60,column=135).value=sz26bpresent
        sheetname.cell(row=61,column=135).value=sz26babsent
        sheetname.cell(row=62,column=135).value=sz26bpassed
        sheetname.cell(row=63,column=135).value=sz26bfailed
        sheetname.cell(row=64,column=135).value=sz26bpassper
        
        sz26cpassper=(sz26cpassed/sz26cpresent)*100 if sz26cpassed and sz26cpresent !=0 else 0

        sheetname.cell(row=60,column=138).value=sz26cpresent
        sheetname.cell(row=61,column=138).value=sz26cabsent
        sheetname.cell(row=62,column=138).value=sz26cpassed
        sheetname.cell(row=63,column=138).value=sz26cfailed
        sheetname.cell(row=64,column=138).value=sz26cpassper
        
        sz261passper=(sz261passed/sz261present)*100 if sz261passed and sz261present !=0 else 0

        sheetname.cell(row=60,column=141).value=sz261present
        sheetname.cell(row=61,column=141).value=sz261absent
        sheetname.cell(row=62,column=141).value=sz261passed
        sheetname.cell(row=63,column=141).value=sz261failed
        sheetname.cell(row=64,column=141).value=sz261passper
        
        su46bpassper=(su46bpassed/su46bpresent)*100 if su46bpassed and su46bpresent !=0 else 0

        sheetname.cell(row=60,column=144).value=su46bpresent
        sheetname.cell(row=61,column=144).value=su46babsent
        sheetname.cell(row=62,column=144).value=su46bpassed
        sheetname.cell(row=63,column=144).value=su46bfailed
        sheetname.cell(row=64,column=144).value=su46bpassper
        
        sz26qpassper=(sz26qpassed/sz26qpresent)*100 if sz26qpassed and sz26qpresent !=0 else 0

        sheetname.cell(row=60,column=147).value=sz26qpresent
        sheetname.cell(row=61,column=147).value=sz26qabsent
        sheetname.cell(row=62,column=147).value=sz26qpassed
        sheetname.cell(row=63,column=147).value=sz26qfailed
        sheetname.cell(row=64,column=147).value=sz26qpassper
        
        ces6qpassper=(ces6qpassed/ces6qpresent)*100 if ces6qpassed and ces6qpresent !=0 else 0

        sheetname.cell(row=60,column=150).value=ces6qpresent
        sheetname.cell(row=61,column=150).value=ces6qabsent
        sheetname.cell(row=62,column=150).value=ces6qpassed
        sheetname.cell(row=63,column=150).value=ces6qfailed
        sheetname.cell(row=64,column=150).value=ces6qpassper
        
        wb.save(fn)
        print("\n\n"+str(print_sheetname)+" "+"SHEET"+" "+"COMPLETED !")
